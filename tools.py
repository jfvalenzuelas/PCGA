from openpyxl import load_workbook
import pandas as pd
import spacy

global utils_file
global nlp

utils_file = 'utils-search.xlsx'
nlp = spacy.load('es')

def cleanText(text):
    text = text.lower().strip()

    if ('.' in text):
        text = text.replace('.', ' ')

    if ('-' in text):
        text = text.replace('-', ' ')

    if ('cuente' in text):
        text = text.replace('cuente', 'cuenta')

    if ('cte' in text):
        text = text.replace('cte', 'corriente')

    if ('impto' in text):
        text = text.replace('impto', 'impuesto')

    if ('cta' in text):
        text = text.replace('cta', 'cuenta')

    if ('vta' in text):
        text = text.replace('vta', 'venta')

    if ('tarj' in text):
        text = text.replace('tarj', 'tarjeta')

    if ('serv' in text):
        text = text.replace('serv', 'servicios')

    if ('indem' in text):
        text = text.replace('indem', 'indemnizacion')

    if ('maq ' in text):
        text = text.replace('maq ', 'maquinas ')

    if ('gtos' in text):
        text = text.replace('gtos', 'gastos')

    return text

def matchAccount(text):
    text = cleanText(text)
    print('Buscando: '+text)
    wb = load_workbook(utils_file)
    sheet = wb['grupos']
    absolut_simils = []
    columns = []

    for j in range (1, 80):
        all_simils = []

        for i in range (1, sheet.max_row):
            print('Row: '+str(i)+' Column: '+str(j))
            current_cell = sheet.cell(row=i, column=j)
            
            if (current_cell.value == None):
                print('Value: None')
                break
            
            else:
                print('Value: '+str(current_cell.value.lower().strip()))
                doc1 = nlp(text)
                doc2 = nlp(cleanText(current_cell.value.lower().strip()))
                simil = doc1.similarity(doc2)
                all_simils.append(simil)

        print(all_simils)
        if (len(all_simils) > 0):
            absolut_simils.append(max(all_simils))
            columns.append(current_cell.column)
        else:
            absolut_simils.append(float(0))
            columns.append(float(0))
    
    group = columns[absolut_simils.index(max(absolut_simils))]
    value = max(absolut_simils)
    print(sheet[group+'1'].value.lower().strip())
    return group, value